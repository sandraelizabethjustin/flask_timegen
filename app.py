from flask import Flask, render_template, request, send_file
import pandas as pd
import random
import xlsxwriter
import openpyxl

classes = []
app = Flask(__name__)
TOTAL_HRS = 7
DAYS = 5
MAX_SIZE = 120
GAP = 17

def populate_teacher(s):
    length = len(s)
    class_ind = {}
    for k in range(length):
        if str(s.iat[k, 1]).lower() != 'nan':
            class_ind[s.iat[k, 0]] = s.iat[k, 1].split(',')
        else:
            classes.append(str(s.iat[k, 0]))
    return class_ind

def populate(s):
    length = len(s)
    list_1 = []
    class_ind = []
    for k in range(length):
        if str(s.iat[k, 1]).lower() != 'nan':
            class_ind.append(list((s.iat[k, 0], s.iat[k, 1])))
        else:
            if class_ind:
                list_1.append(class_ind)
            classes.append(str(s.iat[k, 0]))
            class_ind = []
    list_1.append(class_ind)
    return list_1

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/page2.html')
def page2():
    return render_template('page2.html')

@app.route('/view', methods=['POST'])
def view():
    files = [request.files[f'file{i}'] for i in range(1, 4)]
    
    s2 = pd.read_excel(files[0], skiprows=2)
    s1 = pd.read_excel(files[1])
    s3 = pd.read_excel(files[2])
    
    path = files[0]
    wb_obj = openpyxl.load_workbook(path)   
    sheet_obj = wb_obj.active 
    cell_obj1 = str(sheet_obj.cell(row=1, column=1).value)
    cell_obj2 = str(sheet_obj.cell(row=2, column=1).value)
    
    output_path = '/tmp/final.xlsx'  # Save to /tmp directory
    wb = xlsxwriter.Workbook(output_path)
    ws = wb.add_worksheet("TimeTable")
    ws2 = wb.add_worksheet("TeacherSlot")
    
    f2 = wb.add_format({'bold': True, 'bg_color': '#b2b2b2'})
    f3 = wb.add_format({'bg_color': '#808080'})
    f4 = wb.add_format({'bold': True, 'bg_color': '#808080'})
    f5 = wb.add_format({'bg_color': '#b2b2b2'})
    f6 = wb.add_format({'bold': True, 'bg_color': '#999999'})
    f7 = wb.add_format({'bold': True})
    working_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    teachers = s1['faculty'].dropna().unique().tolist()
    tt = []
    for teacher in teachers:
        a = teacher.split(',')
        if len(a) == 1:
            tt.append(a[0])
        else:
            for i in a:
                tt.append(i)
    teachers = list(set(tt))

    teacher_course = populate_teacher(s1)
    teacher_len = len(teachers)

    t_len = len(s2)
    course_hour = populate(s3)
    timeslot = [[0] * MAX_SIZE for _ in range(t_len)]
    teacherslot = [[0] * MAX_SIZE for _ in range(teacher_len)]

    for i in range(t_len):
        index = 0
        for k in range(DAYS):
            for j in range(TOTAL_HRS):
                timeslot[i][index] = str(s2.iat[i, (k * TOTAL_HRS) + j])
                if timeslot[i][index] in teacher_course:
                    fac = teacher_course[timeslot[i][index]]
                    t_index = [teachers.index(teacher) for teacher in fac]
                    for tindex in t_index:
                         teacherslot[tindex][index] = timeslot[i][index]
                index += 1
            index += GAP

    for k in range(t_len):
        c_h = course_hour[k]
        for course, hour in c_h:
            fac = teacher_course[course]
            t_index = [teachers.index(teacher) for teacher in fac]
            rem_hr = hour
            while int(rem_hr) > 0:
                for j in range(MAX_SIZE):
                    if str(timeslot[k][j]).lower() == "nan":
                        begin = j
                        break
                for j in range((MAX_SIZE - 1), -1, -1):
                    if str(timeslot[k][j]).lower() == "nan":
                        end = j
                        break
                interval = (end - begin + 1) / (rem_hr - 1) if rem_hr != 1 else 1
                slots = [begin + j * interval for j in range(int(rem_hr))]
                for slot in slots:
                    if str(timeslot[k][int(slot)]).lower() == "nan":
                        if all(teacherslot[tindex][int(slot)] == 0 and 
                               teacherslot[tindex][(int(slot) - 1) % MAX_SIZE] == 0 and 
                               teacherslot[tindex][(int(slot) + 1) % MAX_SIZE] == 0 
                               for tindex in t_index):
                            timeslot[k][int(slot)] = course
                            for tindex in t_index:
                                teacherslot[tindex][int(slot)] = course
                            rem_hr -= 1
                            break
                    else:
                        left, right = int(slot) - 1, int(slot) + 1
                        while left > 0 or right < MAX_SIZE:
                            if left > 0 and str(timeslot[k][left]).lower() == "nan":
                                if all(teacherslot[tindex][left] == 0 and 
                                       teacherslot[tindex][(left - 1) % MAX_SIZE] == 0 and 
                                       teacherslot[tindex][(left + 1) % MAX_SIZE] == 0 
                                       for tindex in t_index):
                                    timeslot[k][left] = course
                                    for tindex in t_index:
                                        teacherslot[tindex][left] = course
                                    rem_hr -= 1
                                    break
                            if right < MAX_SIZE and str(timeslot[k][right]).lower() == "nan":
                                if all(teacherslot[tindex][right] == 0 and 
                                       teacherslot[tindex][(right - 1) % MAX_SIZE] == 0 and 
                                       teacherslot[tindex][(right + 1) % MAX_SIZE] == 0 
                                       for tindex in t_index):
                                    timeslot[k][right] = course
                                    for tindex in t_index:
                                        teacherslot[tindex][right] = course
                                    rem_hr -= 1
                                    break
                            left -= 1
                            right += 1
                        if left < 0 and right >= MAX_SIZE:
                            print("ERROR: ALLOCATION COULD NOT BE DONE")
                            break

    k = 0
    timetable = []
    counter = 3
    merge_format = wb.add_format({"bold": 1, "align": "center", "valign": "vcenter"})
    ws.merge_range("A1:U1", cell_obj1, merge_format)
    ws.merge_range("A2:U2", cell_obj2, merge_format)

    while k < t_len:
        index = 0
        temp = [[0] * TOTAL_HRS for _ in range(DAYS)]
        timetable.append(['', '', '', classes[k], '', '', ''])
        for i in range(DAYS):
            for j in range(TOTAL_HRS):
                if str(timeslot[k][index]).lower() == "nan":
                    timeslot[k][index] = "REMEDIAL"
                temp[i][j] = timeslot[k][index]           
                index += 1
            timetable.append(temp[i])
            index += GAP
        ws.write(counter, 4, classes[k])
        if k % 2 == 0:
            ws.write_row(counter + 1, 0, ['', '1st', '2nd', '3rd', 'Lunch', '4th', '5th', '6th'], f6)
        else:
            ws.write_row(counter + 1, 0, ['', '1st', '2nd', '3rd', 'Lunch', '4th', '5th', '6th'], f4)
        for row in range(len(temp)):
            ws.write(counter + row + 2, 0, working_days[row], f6 if row % 2 == 0 else f4)
            for col, value in enumerate(temp[row]):
                ws.write(counter + row + 2, col + 1, value, f5 if row % 2 == 0 else f3)
        counter += 8
        k += 1

    k = 0
    teachslot = []
    counter = 3
    ws2.merge_range("A1:U1", cell_obj1, merge_format)
    ws2.merge_range("A2:U2", cell_obj2, merge_format)

    while k < teacher_len:
        index = 0
        temp = [[0] * TOTAL_HRS for _ in range(DAYS)]
        teachslot.append([teachers[k], '', '', '', '', '', ''])
        for i in range(DAYS):
            for j in range(TOTAL_HRS):
                if str(teacherslot[k][index]).lower() == "nan":
                    teacherslot[k][index] = "FREE"
                temp[i][j] = teacherslot[k][index]           
                index += 1
            teachslot.append(temp[i])
            index += GAP
        ws2.write(counter, 4, teachers[k])
        if k % 2 == 0:
            ws2.write_row(counter + 1, 0, ['', '1st', '2nd', '3rd', 'Lunch', '4th', '5th', '6th'], f6)
        else:
            ws2.write_row(counter + 1, 0, ['', '1st', '2nd', '3rd', 'Lunch', '4th', '5th', '6th'], f4)
        for row in range(len(temp)):
            ws2.write(counter + row + 2, 0, working_days[row], f6 if row % 2 == 0 else f4)
            for col, value in enumerate(temp[row]):
                ws2.write(counter + row + 2, col + 1, value, f5 if row % 2 == 0 else f3)
        counter += 8
        k += 1

    wb.close()
    return send_file(output_path, as_attachment=True)

# if __name__ == '__main__':
#     app.run(debug=True)
